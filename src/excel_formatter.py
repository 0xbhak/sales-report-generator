#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Modul untuk styling worksheet Excel
"""

from openpyxl.styles import Font, Alignment, Border, Side, PatternFill


def apply_styling(ws, data_penjualan, tanggal, total, store_config):
    """
    Apply styling ke worksheet
    
    Args:
        ws: Worksheet object
        data_penjualan (list): List of dict berisi data penjualan
        tanggal (str): Tanggal penjualan
        total (int): Total penjualan
        store_config (StoreConfig): Konfigurasi toko
    """
    
    # Border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Header toko
    ws['A1'] = store_config.nama_toko
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A1:E1')
    
    ws['A2'] = store_config.alamat
    ws['A2'].font = Font(size=10)
    ws['A2'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A2:E2')
    
    ws['A3'] = store_config.telepon
    ws['A3'].font = Font(size=10)
    ws['A3'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A3:E3')
    
    # Tanggal
    ws['A5'] = f'Tanggal : {tanggal}'
    ws['A5'].font = Font(bold=True)
    
    # Judul tabel
    ws['A7'] = 'RINCIAN PENJUALAN'
    ws['A7'].font = Font(bold=True, size=12)
    ws['A7'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A7:E7')
    
    # Header tabel
    headers = ['No', 'Nama Produk', 'Jumlah', 'Harga Satuan', 'Sub Total']
    header_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=8, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Data penjualan
    row_num = 9
    for idx, item in enumerate(data_penjualan, 1):
        ws.cell(row=row_num, column=1, value=idx)
        ws.cell(row=row_num, column=2, value=item['Nama Produk'])
        ws.cell(row=row_num, column=3, value=item['Jumlah'])
        ws.cell(row=row_num, column=4, value=item['Harga Satuan'])
        ws.cell(row=row_num, column=5, value=item['Sub Total'])
        
        # Apply border
        for col in range(1, 6):
            ws.cell(row=row_num, column=col).border = thin_border
            
            # Alignment
            if col == 1 or col == 3:  # No dan Jumlah
                ws.cell(row=row_num, column=col).alignment = Alignment(horizontal='center')
            elif col == 4 or col == 5:  # Harga Satuan dan Sub Total
                ws.cell(row=row_num, column=col).alignment = Alignment(horizontal='right')
                # Format angka dengan pemisah ribuan (format Indonesia: #.##0)
                ws.cell(row=row_num, column=col).number_format = '#,##0'
        
        row_num += 1
    
    # Total
    ws.cell(row=row_num, column=4, value='Jumlah')
    ws.cell(row=row_num, column=4).font = Font(bold=True)
    ws.cell(row=row_num, column=4).fill = header_fill
    ws.cell(row=row_num, column=4).border = thin_border
    ws.cell(row=row_num, column=4).alignment = Alignment(horizontal='center')
    
    ws.cell(row=row_num, column=5, value=total)
    ws.cell(row=row_num, column=5).font = Font(bold=True)
    ws.cell(row=row_num, column=5).fill = header_fill
    ws.cell(row=row_num, column=5).border = thin_border
    ws.cell(row=row_num, column=5).alignment = Alignment(horizontal='right')
    # Format angka total dengan pemisah ribuan
    ws.cell(row=row_num, column=5).number_format = '#,##0'
    
    # Set column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
